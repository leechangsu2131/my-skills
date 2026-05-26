"""
S2B 학교장터 - 구매 결과 Excel 리포트 생성 모듈

견적서 담기 결과를 Excel 파일로 저장합니다.
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =====================================================
# 색상 상수
# =====================================================
COLOR_HEADER_BG    = "1F4E79"   # 진한 파랑 (헤더 배경)
COLOR_HEADER_FG    = "FFFFFF"   # 흰색 (헤더 글자)
COLOR_SUCCESS_BG   = "E2EFDA"   # 연한 초록 (성공 행)
COLOR_FAIL_BG      = "FCE4D6"   # 연한 빨강 (실패 행)
COLOR_SUMMARY_BG   = "D6E4F0"   # 연한 파랑 (요약 행)
COLOR_TITLE_FG     = "1F4E79"   # 진한 파랑 (제목 글자)


def _thin_border():
    """얇은 테두리 스타일 반환"""
    thin = Side(style='thin', color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_header_cell(ws, row, col, value, width=None):
    """헤더 셀 스타일 적용"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="맑은 고딕", bold=True, color=COLOR_HEADER_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()
    return cell


def _make_data_cell(ws, row, col, value, bg_color=None, bold=False, align="left"):
    """데이터 셀 스타일 적용"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="맑은 고딕", bold=bold, size=10)
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _thin_border()
    return cell


def save_report(results: list, output_path: str = None) -> str:
    """
    구매 결과를 Excel 파일로 저장합니다.

    Args:
        results: 결과 딕셔너리 리스트
            [
                {
                    'request_name': '요청 품목명',
                    'quantity': 수량(int),
                    'selected_title': '선택된 물품명',   # 검색 결과 없으면 빈 문자열
                    'selected_id': '물품번호',           # 검색 결과 없으면 빈 문자열
                    'success': True/False,
                    'fail_reason': '실패 사유',          # 성공 시 빈 문자열
                    'processed_at': datetime 객체,
                }
            ]
        output_path: 저장할 파일 경로 (None이면 자동 생성)

    Returns:
        str: 저장된 파일 경로
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl이 설치되지 않았습니다.\n"
            "pip install openpyxl --break-system-packages"
        )

    # 저장 경로 자동 생성
    if output_path is None:
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"s2b_구매결과_{now_str}.xlsx"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "견적서 담기 결과"

    # ── 제목 행 ──────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"S2B 학교장터 견적서 담기 결과  |  {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}"
    title_cell.font = Font(name="맑은 고딕", bold=True, size=13, color=COLOR_TITLE_FG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="EBF3FB")
    ws.row_dimensions[1].height = 28

    # ── 헤더 행 (2행) ────────────────────────────────
    headers = [
        ("No.",          5),
        ("요청 품목명",  28),
        ("수량",          7),
        ("선택된 물품명", 40),
        ("물품번호",      18),
        ("결과",           8),
        ("비고 / 실패 사유", 30),
    ]
    for col_idx, (header, col_width) in enumerate(headers, start=1):
        _make_header_cell(ws, 2, col_idx, header)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 22

    # ── 데이터 행 (3행~) ─────────────────────────────
    success_count = 0
    fail_count = 0

    for row_idx, result in enumerate(results, start=1):
        excel_row = row_idx + 2
        is_success = result.get("success", False)
        bg = COLOR_SUCCESS_BG if is_success else COLOR_FAIL_BG

        processed_at = result.get("processed_at")
        time_str = processed_at.strftime("%H:%M:%S") if processed_at else ""

        _make_data_cell(ws, excel_row, 1, row_idx,          bg, align="center")
        _make_data_cell(ws, excel_row, 2, result.get("request_name", ""))
        _make_data_cell(ws, excel_row, 3, result.get("quantity", ""),   bg, align="center")
        _make_data_cell(ws, excel_row, 4, result.get("selected_title", ""))
        _make_data_cell(ws, excel_row, 5, result.get("selected_id", ""), align="center")

        # 결과 셀 (성공/실패 표시)
        result_cell = _make_data_cell(
            ws, excel_row, 6,
            "✅ 성공" if is_success else "❌ 실패",
            bg, bold=True, align="center"
        )
        result_cell.font = Font(
            name="맑은 고딕", bold=True, size=10,
            color="375623" if is_success else "9C0006"
        )

        # 비고: 성공 시 처리시각, 실패 시 사유
        note = time_str if is_success else result.get("fail_reason", "")
        _make_data_cell(ws, excel_row, 7, note)
        ws.row_dimensions[excel_row].height = 18

        if is_success:
            success_count += 1
        else:
            fail_count += 1

    # ── 요약 행 ──────────────────────────────────────
    summary_row = len(results) + 3
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    summary_label = ws[f"A{summary_row}"]
    summary_label.value = "합계"
    summary_label.font = Font(name="맑은 고딕", bold=True, size=10)
    summary_label.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)
    summary_label.alignment = Alignment(horizontal="right", vertical="center")
    summary_label.border = _thin_border()

    for col in range(2, 6):  # 병합 셀 나머지에도 테두리 적용
        ws.cell(summary_row, col).border = _thin_border()
        ws.cell(summary_row, col).fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)

    summary_result = ws.cell(summary_row, 6,
        f"✅ {success_count}건 / ❌ {fail_count}건"
    )
    summary_result.font = Font(name="맑은 고딕", bold=True, size=10)
    summary_result.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)
    summary_result.alignment = Alignment(horizontal="center", vertical="center")
    summary_result.border = _thin_border()

    summary_note = ws.cell(summary_row, 7,
        f"총 {len(results)}건 처리  |  성공률 {success_count/len(results)*100:.0f}%"
        if results else "처리 결과 없음"
    )
    summary_note.font = Font(name="맑은 고딕", size=10, color="595959")
    summary_note.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)
    summary_note.alignment = Alignment(horizontal="left", vertical="center")
    summary_note.border = _thin_border()
    ws.row_dimensions[summary_row].height = 20

    # 창 틀 고정 (헤더 아래)
    ws.freeze_panes = "A3"

    # 인쇄 설정
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:2"

    wb.save(output_path)
    return output_path
