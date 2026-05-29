"""
Purchase-list readers for S2B buyer automation.

Supports the quote-sheet workbook format used in the list/ folder:
NO | G2B/S2B번호 | 품명 및 규격 | 수량 | 단위 | 단가 | 금액
"""

from __future__ import annotations

import csv
import os
import re
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


HEADER_ALIASES = {
    "no": ("NO", "순번"),
    "code": ("G2B/S2B번호", "G2BS2B번호", "S2B번호", "G2B번호", "물품번호"),
    "name": ("품명및규격", "품명", "품목명", "검색어"),
    "quantity": ("수량", "QTY", "QUANTITY"),
    "unit": ("단위",),
    "unit_price": ("단가", "단  가", "단   가"),
    "amount": ("금액", "금       액", "합계금액"),
}


def _normalize(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s\u3000:/._()\-]+", "", text).upper()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_quantity(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)

    text = str(value).replace(",", "").strip()
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None

    try:
        number = Decimal(match.group(0))
    except InvalidOperation:
        return None
    return int(number) if number == number.to_integral() else float(number)


def _clean_money(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{int(value):,}" if float(value).is_integer() else f"{value:,}"
    return str(value).strip()


def extract_s2b_id(value: Any) -> str:
    """Return the S2B item id from a G2B/S2B mixed code cell."""
    text = _cell_text(value)
    if not text:
        return ""
    digit_runs = re.findall(r"\d{9,}", text)
    candidates = [run for run in digit_runs if len(run) >= 12]
    return (candidates or digit_runs or [""])[-1]


def _match_header_cell(value: Any) -> str | None:
    normalized = _normalize(value)
    if not normalized:
        return None
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if _normalize(alias) in normalized:
                return field
    return None


def _find_header(ws) -> tuple[int, dict[str, int]]:
    max_row = min(ws.max_row, 80)
    best: tuple[int, dict[str, int], int] | None = None

    for row in ws.iter_rows(min_row=1, max_row=max_row):
        mapping: dict[str, int] = {}
        for idx, cell in enumerate(row, start=1):
            field = _match_header_cell(cell.value)
            if field and field not in mapping:
                mapping[field] = idx

        score = sum(field in mapping for field in ("code", "name", "quantity"))
        if score >= 2 and "name" in mapping and "quantity" in mapping:
            return row[0].row, mapping
        if score and (best is None or score > best[2]):
            best = (row[0].row, mapping, score)

    if best:
        row_num, mapping, _ = best
        if "name" in mapping and "quantity" in mapping:
            return row_num, mapping
    raise ValueError("엑셀에서 품명/수량 헤더 행을 찾지 못했습니다.")


def _is_note_or_summary(row_values: list[str]) -> bool:
    joined = " ".join(row_values)
    if not joined.strip():
        return False
    blocked_terms = ("합계", "부가세 포함", "설치비 포함", "전문생산")
    return any(term in joined for term in blocked_terms)


def _row_to_item(ws, row_num: int, columns: dict[str, int]) -> dict[str, Any] | None:
    def value(field: str) -> Any:
        col = columns.get(field)
        return ws.cell(row_num, col).value if col else None

    name = _cell_text(value("name"))
    raw_code = _cell_text(value("code"))
    quantity = _parse_quantity(value("quantity"))

    row_values = [
        _cell_text(ws.cell(row_num, col).value)
        for col in range(1, min(ws.max_column, 8) + 1)
    ]
    if _is_note_or_summary(row_values):
        return None
    if not name or quantity is None:
        return None

    s2b_id = extract_s2b_id(raw_code)
    unit_price = _clean_money(value("unit_price"))

    return {
        "source_row": row_num,
        "no": _cell_text(value("no")),
        "raw_code": raw_code,
        "s2b_id": s2b_id,
        "name": name,
        "quantity": quantity,
        "unit": _cell_text(value("unit")),
        "unit_price": unit_price,
        "amount": _clean_money(value("amount")),
        "search_query": s2b_id or name,
    }


def parse_quote_workbook(source: str | os.PathLike[str] | BinaryIO) -> list[dict[str, Any]]:
    """Parse an S2B quote-style workbook and return purchase items."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 설치되지 않았습니다.")

    wb = load_workbook(source, data_only=True)
    items: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        try:
            header_row, columns = _find_header(ws)
        except ValueError:
            continue

        empty_streak = 0
        for row_num in range(header_row + 1, ws.max_row + 1):
            item = _row_to_item(ws, row_num, columns)
            if item:
                item["sheet"] = ws.title
                items.append(item)
                empty_streak = 0
                continue

            row_text = " ".join(
                _cell_text(ws.cell(row_num, col).value)
                for col in range(1, min(ws.max_column, 8) + 1)
            )
            empty_streak = empty_streak + 1 if not row_text.strip() else 0
            if empty_streak >= 5:
                break

    if not items:
        raise ValueError("엑셀에서 수량이 있는 구매 품목을 찾지 못했습니다.")
    return items


def parse_csv_items(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    items: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        normalized_keys = {_normalize(key): key for key in row.keys() if key}

        def pick(*aliases: str) -> str:
            for alias in aliases:
                key = normalized_keys.get(_normalize(alias))
                if key:
                    return _cell_text(row.get(key))
            return ""

        name = pick("품목명", "품명", "검색어")
        quantity = _parse_quantity(pick("수량", "qty", "quantity"))
        raw_code = pick("G2B/S2B번호", "S2B번호", "물품번호")
        if not name or quantity is None:
            continue

        s2b_id = extract_s2b_id(raw_code)
        items.append({
            "source_row": index,
            "no": str(len(items) + 1),
            "raw_code": raw_code,
            "s2b_id": s2b_id,
            "name": name,
            "quantity": quantity,
            "unit": pick("단위"),
            "unit_price": pick("단가", "가격", "제시금액"),
            "amount": pick("금액"),
            "search_query": s2b_id or name,
        })

    if not items:
        raise ValueError("CSV에서 구매 품목을 찾지 못했습니다.")
    return items


def load_purchase_items(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    ext = os.path.splitext(os.fspath(path))[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return parse_quote_workbook(path)
    if ext == ".csv":
        return parse_csv_items(path)
    raise ValueError("지원하지 않는 파일 형식입니다. .xlsx 또는 .csv 파일을 사용하세요.")
