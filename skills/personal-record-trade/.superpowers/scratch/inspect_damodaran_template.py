from __future__ import annotations

import json
from pathlib import Path
from datetime import date, datetime

from openpyxl import load_workbook

INPUT = Path(r"D:\Downloads\fcffsimpleginzu.xlsx")
OUT = Path(r"C:\Users\lee21\.gemini\antigravity\scratch\my-skills\skills\personal-record-trade\.superpowers\scratch")

TERMS = [
    "revenue",
    "operating",
    "margin",
    "tax",
    "reinvestment",
    "sales to capital",
    "fcff",
    "cost of capital",
    "wacc",
    "terminal",
    "value",
    "equity",
    "debt",
    "cash",
    "roic",
    "roc",
]


def compact(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.replace("\n", " ").strip()
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def cell_ref(cell):
    return f"{cell.parent.title}!{cell.coordinate}"


def main():
    wb_formula = load_workbook(INPUT, data_only=False, read_only=False, keep_links=True)
    wb_values = load_workbook(INPUT, data_only=True, read_only=False, keep_links=True)

    summary = {
        "file": str(INPUT),
        "sheets": [],
        "defined_names": [],
        "term_hits": {},
        "formula_samples": [],
        "comments": [],
    }

    for ws in wb_formula.worksheets:
        summary["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:30]],
            }
        )

    try:
        for name in wb_formula.defined_names:
            summary["defined_names"].append(str(name))
    except Exception as exc:
        summary["defined_names_error"] = str(exc)

    hits = {term: [] for term in TERMS}
    formula_samples = []
    comments = []

    for ws in wb_formula.worksheets:
        value_ws = wb_values[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                text = str(value).lower()
                for term in TERMS:
                    if term in text and len(hits[term]) < 60:
                        hits[term].append(
                            {
                                "cell": cell_ref(cell),
                                "value": compact(value),
                                "computed": compact(value_ws[cell.coordinate].value),
                            }
                        )
                if isinstance(value, str) and value.startswith("=") and len(formula_samples) < 250:
                    formula_samples.append(
                        {
                            "cell": cell_ref(cell),
                            "formula": value,
                            "computed": compact(value_ws[cell.coordinate].value),
                        }
                    )
                if cell.comment is not None and len(comments) < 120:
                    comments.append(
                        {
                            "cell": cell_ref(cell),
                            "text": compact(cell.comment.text)[:500] if cell.comment.text else "",
                            "author": cell.comment.author,
                        }
                    )

    summary["term_hits"] = hits
    summary["formula_samples"] = formula_samples
    summary["comments"] = comments

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "damodaran_openpyxl_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("SHEETS")
    for sheet in summary["sheets"]:
        print(f"- {sheet['name']}: {sheet['max_row']} rows x {sheet['max_column']} cols")
    print("\nKEY HITS")
    for term in ["revenue", "operating", "reinvestment", "sales to capital", "fcff", "cost of capital", "terminal", "equity"]:
        print(f"\n[{term}]")
        for hit in hits[term][:8]:
            print(f"{hit['cell']}: {hit['value']} -> {hit['computed']}")
    print("\nFORMULA SAMPLES")
    for item in formula_samples[:40]:
        print(f"{item['cell']}: {item['formula']} -> {item['computed']}")


if __name__ == "__main__":
    main()
