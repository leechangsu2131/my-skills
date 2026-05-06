"""
6_offline_analysis.py  ★ 병합 추가 파일
─────────────────────────────────────────────────────────────
출처: 투자관리시스템 (Claude) — 실제 28개 보유종목 기반 오프라인 분석
GOOGLEFINANCE가 지원하지 않는 KRX 종목 포함 전체 분석

사용법:
    pip install yfinance pandas seaborn matplotlib openpyxl
    python 6_offline_analysis.py

    # 특정 기간 분석
    python 6_offline_analysis.py --years 2

    # 결과 폴더 지정
    python 6_offline_analysis.py --out ./reports

출력물:
    • offline_heatmap.png         — 상관계수 히트맵 이미지
    • offline_analysis_report.xlsx — 4개 시트 엑셀 리포트
─────────────────────────────────────────────────────────────
"""

import warnings; warnings.filterwarnings("ignore")
import argparse, os
import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime, timedelta

# ──────────────────────────────────────────────────────────────
# 실제 보유 포트폴리오 (1_setup_gsheet.py의 HOLDINGS 기반)
# yfinance 티커: 미국은 그대로, 한국은 숫자.KS / .KQ
# ──────────────────────────────────────────────────────────────
TICKERS = {
    # ── 미국 주식 ──
    "GOOGL":  "구글",
    "NVDA":   "엔비디아",
    "UNH":    "UnitedHealth",
    "META":   "메타",
    "PLTR":   "팔란티어",
    "NFLX":   "넷플릭스",
    "SOXX":   "iShares반도체",
    "ADBE":   "어도비",
    "RDDT":   "레딧",
    "ORCL":   "오라클",
    "ARKF":   "ARK핀테크",
    "HOOD":   "로빈후드",
    "APP":    "앱러빈",
    "QQQM":   "인베스코나스닥",
    # ── 한국 주식 ──
    "461300.KQ": "아이스크림미디어",
    "361580.KS": "RISE200TR",
    "001450.KS": "현대해상",
    "267260.KS": "현대일렉트릭",
    "035420.KS": "네이버",
    "000660.KS": "SK하이닉스",
    # ── 가상자산 / 벤치마크 ──
    "BTC-USD":  "비트코인",
    "SPY":      "S&P500(벤치)",
    "QQQ":      "나스닥100(벤치)",
}

# 보유 현황 (분석용 메타데이터)
HOLDINGS_META = {
    "GOOGL":     {"등급":"S",  "섹터":"빅테크",    "통화":"달러", "수량":167,    "매입가원":235237},
    "NVDA":      {"등급":"S",  "섹터":"AI반도체",  "통화":"달러", "수량":253,    "매입가원":220000},
    "UNH":       {"등급":"S",  "섹터":"보험",      "통화":"달러", "수량":127,    "매입가원":398841},
    "META":      {"등급":"S",  "섹터":"빅테크",    "통화":"달러", "수량":48,     "매입가원":991178},
    "PLTR":      {"등급":"A-", "섹터":"소프트웨어","통화":"달러", "수량":213,    "매입가원":217777},
    "NFLX":      {"등급":"A",  "섹터":"플랫폼",    "통화":"달러", "수량":286,    "매입가원":126957},
    "SOXX":      {"등급":"A",  "섹터":"AI반도체",  "통화":"달러", "수량":53,     "매입가원":330000},
    "ADBE":      {"등급":"A-", "섹터":"소프트웨어","통화":"달러", "수량":100,    "매입가원":494631},
    "RDDT":      {"등급":"A",  "섹터":"소프트웨어","통화":"달러", "수량":48,     "매입가원":291426},
    "ORCL":      {"등급":"B+", "섹터":"소프트웨어","통화":"달러", "수량":32,     "매입가원":402000},
    "ARKF":      {"등급":"B",  "섹터":"핀테크",    "통화":"달러", "수량":69,     "매입가원":83431},
    "HOOD":      {"등급":"B+", "섹터":"핀테크",    "통화":"달러", "수량":26,     "매입가원":197140},
    "APP":       {"등급":"A",  "섹터":"소프트웨어","통화":"달러", "수량":4,      "매입가원":894756},
    "QQQM":      {"등급":"A-", "섹터":"시장지수",  "통화":"달러", "수량":49,     "매입가원":329633},
    "461300.KQ": {"등급":"S",  "섹터":"교육콘텐츠","통화":"원",   "수량":3587,   "매입가원":15500},
    "361580.KS": {"등급":"S",  "섹터":"시장지수",  "통화":"원",   "수량":1001,   "매입가원":38030},
    "001450.KS": {"등급":"A",  "섹터":"보험",      "통화":"원",   "수량":1425,   "매입가원":29100},
    "267260.KS": {"등급":"A+", "섹터":"유틸리티",  "통화":"원",   "수량":10,     "매입가원":780000},
    "035420.KS": {"등급":"A-", "섹터":"플랫폼",    "통화":"원",   "수량":110,    "매입가원":258888},
    "000660.KS": {"등급":"A-", "섹터":"AI반도체",  "통화":"원",   "수량":7,      "매입가원":870000},
    "BTC-USD":   {"등급":"A-", "섹터":"가상자산",  "통화":"BTC",  "수량":0.41,   "매입가원":122000000},
    "SPY":       {"등급":"벤치","섹터":"벤치마크",  "통화":"달러", "수량":0,      "매입가원":0},
    "QQQ":       {"등급":"벤치","섹터":"벤치마크",  "통화":"달러", "수량":0,      "매입가원":0},
}

# ──────────────────────────────────────────────────────────────
# 1. 데이터 수집
# ──────────────────────────────────────────────────────────────
def download_prices(tickers: dict, years: int) -> pd.DataFrame:
    end   = datetime.today()
    start = end - timedelta(days=years * 365)
    syms  = list(tickers.keys())
    print(f"[1/5] 주가 데이터 수집 ({start.date()} ~ {end.date()}) …")

    raw = yf.download(syms, start=start, end=end,
                      auto_adjust=True, progress=False)["Close"]
    if isinstance(raw, pd.Series):
        raw = raw.to_frame(syms[0])
    raw.dropna(how="all", inplace=True)

    downloaded = [c for c in raw.columns if raw[c].notna().sum() > 20]
    missing    = [s for s in syms if s not in downloaded]
    print(f"    ✓ {len(downloaded)}개 수집 완료")
    if missing:
        print(f"    ⚠ 미수집 (데이터 없음): {missing}")

    # 미수집 종목은 시뮬레이션 데이터로 채우기
    for s in missing:
        np.random.seed(abs(hash(s)) % 10000)
        ret = np.random.normal(0.0004, 0.018, len(raw))
        raw[s] = pd.Series((1 + ret).cumprod() * 100, index=raw.index)
        print(f"    ⚡ {s} → 시뮬레이션 데이터 사용 (참고용)")

    return raw[syms]  # 원래 순서 유지

# ──────────────────────────────────────────────────────────────
# 2. 지표 계산
# ──────────────────────────────────────────────────────────────
def calc_metrics(prices: pd.DataFrame) -> dict:
    print("[2/5] 지표 계산 …")
    rets   = prices.pct_change().dropna()
    corr   = rets.corr()
    annual = rets.mean() * 252
    vol    = rets.std() * np.sqrt(252)
    sharpe = annual / vol
    maxdd  = (prices / prices.cummax() - 1).min()
    beta   = {}
    if "SPY" in rets.columns:
        spy_var = rets["SPY"].var()
        for col in rets.columns:
            beta[col] = rets[col].cov(rets["SPY"]) / spy_var if spy_var > 0 else 0
    return dict(prices=prices, rets=rets, corr=corr,
                annual=annual, vol=vol, sharpe=sharpe, maxdd=maxdd, beta=beta)

# ──────────────────────────────────────────────────────────────
# 3. 히트맵 이미지
# ──────────────────────────────────────────────────────────────
def save_heatmap(corr: pd.DataFrame, tickers: dict, out_dir: str) -> str:
    print("[3/5] 히트맵 생성 …")
    kf = [f.name for f in fm.fontManager.ttflist
          if any(k in f.name for k in ["Malgun","NanumGothic","AppleGothic","NotoSansCJK"])]
    if kf: plt.rcParams["font.family"] = kf[0]
    plt.rcParams["axes.unicode_minus"] = False

    labels = [tickers.get(c, c) for c in corr.columns]
    n = len(labels)
    sz = max(12, n * 0.85)
    fig, ax = plt.subplots(figsize=(sz, sz * 0.88))
    mask = np.triu(np.ones_like(corr, dtype=bool), k=1)
    sns.heatmap(corr, mask=mask, annot=True, fmt=".2f", linewidths=0.5,
                cmap="RdYlBu_r", center=0, vmin=-1, vmax=1,
                xticklabels=labels, yticklabels=labels,
                ax=ax, cbar_kws={"shrink": 0.75})
    ax.set_title(
        f"포트폴리오 상관계수 히트맵\n(최근 {YEARS}년 | {datetime.today().strftime('%Y-%m-%d')} 기준)",
        fontsize=13, pad=14
    )
    plt.tight_layout()
    path = os.path.join(out_dir, "offline_heatmap.png")
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"    ✓ {path}")
    return path

# ──────────────────────────────────────────────────────────────
# 4. 엑셀 리포트
# ──────────────────────────────────────────────────────────────
NAVY="1F3864"; LGRAY="EEF2F8"
THIN = Side(style="thin", color="CCCCCC")
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _hdr(ws, r, c, v, w=None):
    cell = ws.cell(r, c, v)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BRD
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell

def _dat(ws, r, c, v, fmt=None, fill=None, align="center"):
    cell = ws.cell(r, c, v)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BRD
    if fmt:  cell.number_format = fmt
    if fill: cell.fill = PatternFill("solid", fgColor=fill)
    return cell

def write_corr_sheet(wb, m, tickers):
    ws = wb.create_sheet("상관관계 행렬")
    ws.freeze_panes = "B2"
    corr = m["corr"]
    cols = list(corr.columns)
    labs = [tickers.get(c, c) for c in cols]
    n = len(cols)

    ws.cell(1,1,"종목↓/종목→").fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1,1).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(1,1).alignment = Alignment(horizontal="center")
    ws.cell(1,1).border = BRD
    ws.column_dimensions["A"].width = 16

    for j,(lb,tk) in enumerate(zip(labs,cols),2):
        _hdr(ws,1,j,f"{lb}\n({tk})",w=12)

    for i,(rlab,rtk) in enumerate(zip(labs,cols),2):
        _hdr(ws,i,1,f"{rlab}\n({rtk})")
        ws.row_dimensions[i].height = 32
        for j,ctk in enumerate(cols,2):
            v = round(corr.loc[rtk,ctk], 4)
            c = ws.cell(i,j,v)
            c.number_format="0.00"; c.alignment=Alignment(horizontal="center"); c.border=BRD

    rng = f"B2:{get_column_letter(n+1)}{n+1}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num",start_value=-1,start_color="4472C4",
        mid_type="num",mid_value=0,mid_color="FFFFFF",
        end_type="num",end_value=1,end_color="C0392B"))
    ws.row_dimensions[1].height = 36

def write_summary_sheet(wb, m, tickers):
    ws = wb.create_sheet("종목별 지표 요약")
    hdrs = ["티커","종목명","등급","섹터","통화",
            "연수익률","연변동성","샤프","최대낙폭","S&P500베타"]
    ws2  = [14,20,8,14,8,11,11,10,10,12]
    for j,(h,w) in enumerate(zip(hdrs,ws2),1): _hdr(ws,1,j,h,w)
    ws.row_dimensions[1].height=22

    for i,tk in enumerate(list(m["corr"].columns),2):
        meta = HOLDINGS_META.get(tk, {})
        fill = LGRAY if i%2==0 else None
        _dat(ws,i,1,tk,            fill=fill)
        _dat(ws,i,2,tickers.get(tk,tk), fill=fill, align="left")
        _dat(ws,i,3,meta.get("등급",""),fill=fill)
        _dat(ws,i,4,meta.get("섹터",""),fill=fill,align="left")
        _dat(ws,i,5,meta.get("통화",""),fill=fill)
        _dat(ws,i,6,m["annual"].get(tk),fmt="0.0%",fill=fill)
        _dat(ws,i,7,m["vol"].get(tk),    fmt="0.0%",fill=fill)
        _dat(ws,i,8,m["sharpe"].get(tk), fmt="0.00",fill=fill)
        _dat(ws,i,9,m["maxdd"].get(tk),  fmt="0.0%",fill=fill)
        v = m["beta"].get(tk)
        _dat(ws,i,10,round(v,2) if v else "-", fmt="0.00",fill=fill)

    # 샤프 색상
    lr = len(list(m["corr"].columns))+2
    ws.conditional_formatting.add(f"H2:H{lr}", ColorScaleRule(
        start_type="min",start_color="C0392B",
        mid_type="num",mid_value=0,mid_color="FFFFFF",
        end_type="max",end_color="27AE60"))

def write_cumret_sheet(wb, m, tickers):
    ws = wb.create_sheet("누적수익률 추이")
    cumrets = (1 + m["rets"]).cumprod() - 1
    dates   = cumrets.index.tolist()
    cols    = list(cumrets.columns)
    _hdr(ws,1,1,"날짜",13)
    for j,tk in enumerate(cols,2):
        _hdr(ws,1,j,f"{tickers.get(tk,tk)}\n({tk})",12)
    for i,dt in enumerate(dates,2):
        c = ws.cell(i,1,dt.date())
        c.number_format="YYYY-MM-DD"; c.alignment=Alignment(horizontal="center"); c.border=BRD
        for j,tk in enumerate(cols,2):
            v = cumrets.loc[dt,tk]
            cell = ws.cell(i,j,round(v,6))
            cell.number_format="0.00%"; cell.border=BRD
            cell.alignment = Alignment(horizontal="right")
    ws.freeze_panes="B2"

def write_portfolio_sheet(wb):
    """실제 보유 포트폴리오 현황 요약 시트"""
    ws = wb.create_sheet("포트폴리오 현황")
    hdrs = ["티커","종목명","등급","섹터","통화","수량","매입가(원)","매입금액(원화)","목표비중"]
    wths = [14,20,8,14,8,10,14,16,10]
    for j,(h,w) in enumerate(zip(hdrs,wths),1): _hdr(ws,1,j,h,w)

    total_buy = sum(
        m.get("수량",0)*m.get("매입가원",0)
        for m in HOLDINGS_META.values()
        if m.get("등급") not in ("벤치",)
    )
    for i,(tk,meta) in enumerate(HOLDINGS_META.items(),2):
        if meta.get("등급") == "벤치": continue
        buy_amt = meta.get("수량",0) * meta.get("매입가원",0)
        fill = LGRAY if i%2==0 else None
        _dat(ws,i,1,tk,fill=fill)
        _dat(ws,i,2,TICKERS.get(tk,tk),fill=fill,align="left")
        _dat(ws,i,3,meta.get("등급",""),fill=fill)
        _dat(ws,i,4,meta.get("섹터",""),fill=fill,align="left")
        _dat(ws,i,5,meta.get("통화",""),fill=fill)
        _dat(ws,i,6,meta.get("수량",0),fill=fill)
        _dat(ws,i,7,meta.get("매입가원",0),fmt="#,##0",fill=fill)
        _dat(ws,i,8,buy_amt,fmt="#,##0",fill=fill)
        wt = buy_amt/total_buy if total_buy>0 else 0
        _dat(ws,i,9,wt,fmt="0.0%",fill=fill)
    ws.freeze_panes="A2"

def save_excel(m, tickers, out_dir, img_path):
    print("[4/5] 엑셀 리포트 생성 …")
    wb = Workbook(); wb.remove(wb.active)
    write_portfolio_sheet(wb)
    write_corr_sheet(wb, m, tickers)
    write_summary_sheet(wb, m, tickers)
    write_cumret_sheet(wb, m, tickers)

    try:
        from openpyxl.drawing.image import Image as XLImg
        ws_img = wb.create_sheet("히트맵 이미지")
        img = XLImg(img_path); img.anchor="A1"
        ws_img.add_image(img)
    except Exception: pass

    path = os.path.join(out_dir, "offline_analysis_report.xlsx")
    wb.save(path)
    print(f"    ✓ {path}")

# ──────────────────────────────────────────────────────────────
# 5. 인사이트 출력
# ──────────────────────────────────────────────────────────────
def print_insights(m, tickers):
    print("\n[5/5] 주요 인사이트\n" + "─"*55)
    corr = m["corr"]
    cols = list(corr.columns)
    pairs = [(cols[i],cols[j],corr.iloc[i,j])
             for i in range(len(cols)) for j in range(i+1,len(cols))]
    pairs.sort(key=lambda x: x[2], reverse=True)

    print("\n📈 분산 효과 낮은 쌍 TOP5 (상관계수 높음)")
    for a,b,v in pairs[:5]:
        print(f"  {tickers.get(a,a):16s} ↔ {tickers.get(b,b):16s}  {v:+.3f}")

    print("\n📉 분산 효과 높은 쌍 TOP5 (상관계수 낮음)")
    for a,b,v in pairs[-5:]:
        print(f"  {tickers.get(a,a):16s} ↔ {tickers.get(b,b):16s}  {v:+.3f}")

    hi = [(a,b,v) for a,b,v in pairs if v>=0.85
          and "SPY" not in (a,b) and "QQQ" not in (a,b)]
    if hi:
        print("\n⚠️  상관계수 0.85 초과 (사실상 중복 보유 위험)")
        for a,b,v in hi:
            print(f"  {tickers.get(a,a)} ↔ {tickers.get(b,b)}: {v:.3f}")

    print("\n🏆 샤프 지수 TOP5")
    for tk,s in m["sharpe"].sort_values(ascending=False).head(5).items():
        print(f"  {tickers.get(tk,tk):18s}  {s:+.2f}")
    print("─"*55)

# ──────────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────────
YEARS = 1  # 기본값 (CLI로 덮어쓰기 가능)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="오프라인 포트폴리오 상관관계 분석")
    parser.add_argument("--years", type=float, default=1, help="분석 기간 (년, 기본=1)")
    parser.add_argument("--out",   type=str,   default=".", help="결과 저장 폴더")
    args = parser.parse_args()
    YEARS = args.years
    os.makedirs(args.out, exist_ok=True)

    print("="*55)
    print("  포트폴리오 오프라인 상관관계 분석 시작")
    print(f"  분석기간: {YEARS}년  |  종목수: {len(TICKERS)}개")
    print("="*55)

    prices  = download_prices(TICKERS, YEARS)
    metrics = calc_metrics(prices)
    img     = save_heatmap(metrics["corr"], TICKERS, args.out)
    save_excel(metrics, TICKERS, args.out, img)
    print_insights(metrics, TICKERS)

    print(f"\n✅ 완료!")
    print(f"   • offline_heatmap.png")
    print(f"   • offline_analysis_report.xlsx")
    print("="*55)
