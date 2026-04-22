"""
학교 전체 안내장 반별 자동 인쇄 (간지 자동 삽입 버전)
=========================================================
동작 흐름:
  ┌─────────────────────────────────────┐
  │ [간지] 3학년 1반 · 17명             │ ← 간지.hwpx 템플릿에서 자동 생성
  │  안내장 × 17장                       │
  │ [간지] 3학년 2반 · 18명             │
  │  안내장 × 18장                       │
  │  ...                                 │
  └─────────────────────────────────────┘
  → 교대 배출 없이도 간지만으로 반 구분 가능
  → 교대 배출까지 켜면 뭉치 경계가 더 명확

준비물:
  - Python 3.8+  (Windows 전용)
  - pip install pywin32 openpyxl
  - 한글 오피스 설치
  - 간지 템플릿 HWPX (SEPARATOR_TEMPLATE 경로)
  - 엑셀 명렬표 (학년/반/학생수 컬럼)

간지 템플릿 만들기:
  한글에서 새 문서 → 아래 두 줄만 크게 작성 후 HWPX 저장
  ─────────────────────────────
  {{학년}}학년 {{반}}반
  총 {{학생수}}명
  ─────────────────────────────
  (폰트 크기를 키워두면 한눈에 잘 보임)
"""

import csv
import os
import re
import sys
import time
import logging
import zipfile
import tempfile
from io import StringIO
from urllib.request import urlopen
from pathlib import Path

try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit("❌ python-dotenv 없음 → pip install python-dotenv")

# ─────────────────────────────────────────
#  설정은 같은 폴더의 .env 파일에서 읽습니다
# ─────────────────────────────────────────
load_dotenv(Path(__file__).parent / ".env")

HWPX_DIR           = os.getenv("HWPX_DIR", "")
EXCEL_FILE         = os.getenv("EXCEL_FILE", "")
SEPARATOR_TEMPLATE = os.getenv("SEPARATOR_TEMPLATE", "")
GOOGLE_SHEET_CSV_URL = os.getenv("GOOGLE_SHEET_CSV_URL", "")

SHEET_NAME  = os.getenv("SHEET_NAME",  "Sheet1")
COL_GRADE   = os.getenv("COL_GRADE",  "학년")
COL_CLASS   = os.getenv("COL_CLASS",  "반")
COL_COUNT   = os.getenv("COL_COUNT",  "학생수")

PRINTER_NAME    = os.getenv("PRINTER_NAME", "")       # 비워두면 기본 프린터
JOB_DELAY       = int(os.getenv("JOB_DELAY", "3"))    # 인쇄 Job 사이 대기(초)
DRY_RUN         = os.getenv("DRY_RUN", "false").lower() == "true"
# 양면 인쇄: 0=단면, 1=양면(긴면/책형), 2=양면(짧은면/달력형) — 간지는 항상 0
DUPLEX          = int(os.getenv("DUPLEX", "0"))
# 간지 용지함: 0=기본값, 1=트레이1(A), 2=트레이2(B), 3=트레이3 ...
SEPARATOR_TRAY  = int(os.getenv("SEPARATOR_TRAY", "0"))

# ─────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────
# 간지 HWPX 동적 생성
# ──────────────────────────────────────────────────────────

def make_separator_hwpx(template_path: str, grade: int, cls: int, count: int) -> str:
    """
    템플릿 HWPX의 {{학년}} {{반}} {{학생수}} 플레이스홀더를
    실제 값으로 교체한 임시 HWPX 파일을 생성하여 경로 반환.
    HWPX는 ZIP이므로 내부 XML을 직접 수정.
    """
    tmp = tempfile.NamedTemporaryFile(suffix=".hwpx", delete=False)
    tmp.close()
    tmp_path = tmp.name

    replacements = {
        "{{학년}}": str(grade),
        "{{반}}": str(cls),
        "{{학생수}}": str(count),
        # 혹시 띄어쓰기 변형이 있을 경우를 위한 대체 패턴
        "{{ 학년 }}": str(grade),
        "{{ 반 }}": str(cls),
        "{{ 학생수 }}": str(count),
    }

    with zipfile.ZipFile(template_path, "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:

        # mimetype은 ZIP_STORED + 첫 번째 엔트리 (HWPX 규격)
        if "mimetype" in zin.namelist():
            zout.writestr(
                zipfile.ZipInfo("mimetype"),  # compression 기본값 = STORED
                zin.read("mimetype")
            )

        for item in zin.infolist():
            if item.filename == "mimetype":
                continue  # 이미 처리함

            data = zin.read(item.filename)

            # XML 파일만 텍스트 치환
            if item.filename.endswith((".xml", ".hpf", ".rels")):
                text = data.decode("utf-8", errors="replace")
                for old, new in replacements.items():
                    text = text.replace(old, new)
                data = text.encode("utf-8")

            zout.writestr(item, data)

    return tmp_path


# ──────────────────────────────────────────────────────────
# 엑셀 로드
# ──────────────────────────────────────────────────────────

def _normalize_header(value: str) -> str:
    return str(value).replace(" ", "").strip()


def _build_class_row(grade, cls, count) -> dict:
    return {
        "grade": int(grade),
        "cls": int(cls),
        "count": int(count),
        "label": f"{int(grade)}학년 {int(cls)}반",
    }


def _parse_rows_with_headers(rows: list[list]) -> list[dict]:
    if not rows:
        return []

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    norm_headers = [_normalize_header(h) for h in headers]

    try:
        gi = norm_headers.index(_normalize_header(COL_GRADE))
        ci = norm_headers.index(_normalize_header(COL_CLASS))
        ni = norm_headers.index(_normalize_header(COL_COUNT))
    except ValueError as e:
        raise ValueError(f"컬럼 없음: {e}  실제 컬럼: {headers}") from e

    classes = []
    for row in rows[1:]:
        if ni >= len(row) or row[ni] in (None, ""):
            continue
        classes.append(_build_class_row(row[gi], row[ci], row[ni]))
    return classes


def _extract_grade_class_count_from_grid(rows: list[list[str]]) -> list[dict]:
    """
    헤더형 테이블이 아닐 때, 셀 전체에서 '1학년', '2반', 숫자 조합을 찾아
    학년/반/학생수를 추출하는 보조 파서.
    """
    classes = []
    pattern_grade = re.compile(r"^\s*(\d+)\s*학년\s*$")
    pattern_class = re.compile(r"^\s*(\d+)\s*반\s*$")
    pattern_number = re.compile(r"^\s*(\d+)\s*$")

    for row in rows:
        values = [str(v).strip() for v in row if v not in (None, "")]
        if len(values) < 3:
            continue
        for i in range(len(values) - 2):
            m_grade = pattern_grade.match(values[i])
            m_class = pattern_class.match(values[i + 1])
            m_count = pattern_number.match(values[i + 2])
            if m_grade and m_class and m_count:
                classes.append(
                    _build_class_row(
                        int(m_grade.group(1)),
                        int(m_class.group(1)),
                        int(m_count.group(1)),
                    )
                )

    dedup = {}
    for c in classes:
        dedup[(c["grade"], c["cls"])] = c
    return [dedup[k] for k in sorted(dedup.keys())]


def load_class_list_from_excel(excel_path: str, sheet: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("❌ openpyxl 없음 → pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    try:
        return _parse_rows_with_headers(rows)
    except ValueError:
        # 학교 시트처럼 구조가 복잡한 경우를 대비해 전체 셀에서 패턴 추출
        parsed = _extract_grade_class_count_from_grid(rows)
        if parsed:
            return parsed
        raise


def load_class_list_from_google_csv(csv_url: str) -> list[dict]:
    with urlopen(csv_url, timeout=15) as response:
        text = response.read().decode("utf-8-sig", errors="replace")

    rows = list(csv.reader(StringIO(text)))
    try:
        return _parse_rows_with_headers(rows)
    except ValueError:
        parsed = _extract_grade_class_count_from_grid(rows)
        if parsed:
            return parsed
        raise


# ──────────────────────────────────────────────────────────
# HWP OLE 헬퍼
# ──────────────────────────────────────────────────────────

def get_hwp():
    try:
        import win32com.client
    except ImportError:
        sys.exit("❌ pywin32 없음 → pip install pywin32")

    try:
        hwp = win32com.client.Dispatch("HWPFrame.HwpObject")
        try:
            hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        except Exception:
            pass
        return hwp
    except Exception as e:
        sys.exit(f"❌ 한글 OLE 연결 실패: {e}")


def _set_printer_bin(printer_name: str, bin_num: int) -> int:
    """
    win32print DEVMODE.DefaultSource 로 프린터 기본 용지함을 임시 변경.
    원래 용지함 번호를 반환 (복원용).
    - GetPrinter(9): 현재 사용자의 기본 프린터 설정 (관리자 권한 불필요)
    """
    if not printer_name or not bin_num:
        return 0
    try:
        import win32print
        h = win32print.OpenPrinter(printer_name)
        info = win32print.GetPrinter(h, 9)          # level 9 = 사용자 권한
        dm = info.get("pDevMode")
        if dm is None:
            win32print.ClosePrinter(h)
            return 0
        orig = dm.DefaultSource
        dm.DefaultSource = bin_num
        win32print.SetPrinter(h, 9, info, 0)
        win32print.ClosePrinter(h)
        log.debug(f"  트레이 DefaultSource 변경 성공: {orig} → {bin_num}")
        return orig
    except Exception as e:
        log.warning(f"  트레이 변경 실패: {e}")
        return 0


def print_hwpx(hwp, file_path: str, copies: int, printer: str,
               duplex: int = 0, paper_source: int = 0) -> bool:
    """안내장 HWPX를 HWP OLE로 열고 copies매 인쇄 후 닫음"""
    orig_bin = 0
    if paper_source and printer:
        orig_bin = _set_printer_bin(printer, paper_source)
        time.sleep(1) # 프린터 설정 시스템 반영 대기

    try:
        hwp.Open(file_path, "HWPX", "forceopen:true")
        time.sleep(1.5)

        act  = hwp.CreateAction("Print")
        hset = act.CreateSet()
        act.GetDefault(hset)

        hset.SetItem("NumCopy", copies)
        hset.SetItem("Range", 0)
        hset.SetItem("Collate", True)
        if printer:
            hset.SetItem("UsePrinterName", printer)
        if duplex:
            hset.SetItem("Duplex", duplex)

        result = act.Execute(hset)
        hwp.Clear(1)
        return bool(result)
    except Exception as e:
        log.error(f"  인쇄 오류: {e}")
        try:
            hwp.Clear(1)
        except Exception:
            pass
        return False
    finally:
        if orig_bin and printer:
            _set_printer_bin(printer, orig_bin)




# ──────────────────────────────────────────────────────────
# 안내장 파일 선택 UI
# ──────────────────────────────────────────────────────────

def select_hwpx_file(hwpx_dir: str) -> str:
    """
    hwpx_dir 에서 *안내장.hwpx 패턴 파일을 찾아 번호 목록을 출력하고
    사용자가 번호를 입력하면 해당 경로를 반환.
    """
    search_dir = Path(hwpx_dir)
    if not search_dir.is_dir():
        sys.exit(f"❌ HWPX_DIR 폴더 없음: {hwpx_dir}")

    candidates = sorted(search_dir.glob("*안내장.hwpx"))
    if not candidates:
        sys.exit(f"❌ '{hwpx_dir}' 에서 *안내장.hwpx 파일을 찾을 수 없습니다.")

    print()
    print("━" * 50)
    print("  인쇄할 안내장 파일을 선택하세요")
    print("━" * 50)
    for i, p in enumerate(candidates, 1):
        print(f"  [{i}] {p.name}")
    print()

    while True:
        raw = input(f"번호 입력 (1~{len(candidates)}): ").strip()
        if raw.isdigit():
            idx = int(raw)
            if 1 <= idx <= len(candidates):
                chosen = candidates[idx - 1]
                print(f"  ✔ 선택됨: {chosen.name}")
                print()
                return str(chosen)
        print(f"  1~{len(candidates)} 사이의 번호를 입력하세요.")
    raise SystemExit("선택 오류")  # unreachable, linter용

def filter_classes_ui(classes):
    """로드된 학급 목록 중 인쇄할 대상을 선택하는 UI"""
    print("\n" + "━" * 50)
    print("  인쇄 범위 선택")
    print("━" * 50)
    print("  [1] 전체 인쇄 (기본)")
    print("  [2] 특정 학년 선택 (예: 4,5,6)")
    print("  [3] 특정 반 부분 검색 (예: 4학년 1반)")
    
    while True:
        sel = input("번호 입력 (1~3, 기본 1): ").strip() or "1"
        if sel == "1":
            return classes
        elif sel == "2":
            g_input = input("인쇄할 학년을 쉼표(,)로 구분해 입력 (예: 4,5,6): ").strip()
            target_grades = [g.strip() for g in g_input.split(",") if g.strip()]
            filtered = [c for c in classes if str(c["grade"]) in target_grades]
            if not filtered:
                print("⚠️ 해당 학년이 없습니다. 다시 선택하세요.")
                continue
            return filtered
        elif sel == "3":
            q = input("검색어 입력 (예: 4학년 1반, 또는 4-1 등 엑셀 표기 기준): ").strip()
            filtered = [c for c in classes if q in c["label"]]
            if not filtered:
                print("⚠️ 검색 결과가 없습니다. 다시 검색하세요.")
                continue
            print("\n[검색 결과]")
            for idx, c in enumerate(filtered, 1):
                print(f"  {idx}. {c['label']} ({c['count']}명)")
            ok = input("위 반들을 인쇄할까요? (Y/N, 기본 Y): ").strip().upper() or "Y"
            if ok == "Y":
                return filtered
            else:
                continue
        else:
            print("⚠️ 잘못된 입력입니다.")


# ──────────────────────────────────────────────────────────
# 메인 루프
# ──────────────────────────────────────────────────────────

def main():
    # ── 안내장 파일 선택 ──────────────────────────────────────
    hwpx_file = select_hwpx_file(HWPX_DIR)

    # ── 나머지 파일 확인 ─────────────────────────────────────
    if not Path(SEPARATOR_TEMPLATE).exists():
        sys.exit(f"❌ 간지 템플릿 HWPX 파일 없음: {SEPARATOR_TEMPLATE}")

    if not GOOGLE_SHEET_CSV_URL and not Path(EXCEL_FILE).exists():
        sys.exit(f"❌ 학생 명렬표 엑셀 파일 없음: {EXCEL_FILE}")

    classes = []
    if GOOGLE_SHEET_CSV_URL:
        log.info("Google Sheets(CSV)에서 반 데이터 로드 시도...")
        try:
            classes = load_class_list_from_google_csv(GOOGLE_SHEET_CSV_URL)
            log.info("Google Sheets 로드 성공")
        except Exception as e:
            log.warning(f"Google Sheets 로드 실패: {e}")
            if Path(EXCEL_FILE).exists():
                log.info("로컬 엑셀 파일로 대체 로드 시도...")
                classes = load_class_list_from_excel(EXCEL_FILE, SHEET_NAME)
            else:
                sys.exit("❌ Google Sheets/엑셀 모두 로드 실패")
    else:
        classes = load_class_list_from_excel(EXCEL_FILE, SHEET_NAME)
    if not classes:
        sys.exit("❌ 엑셀에서 반 데이터를 읽지 못했습니다.")

    classes = filter_classes_ui(classes)

    total = sum(c["count"] for c in classes)
    log.info(f"총 {len(classes)}개 반 / 전체 {total}매 (+ 간지 {len(classes)}장) 예정")
    print()

    # DRY RUN
    if DRY_RUN:
        print("[DRY RUN] 실제 인쇄 안 함 — 인쇄 순서 미리보기:")
        for c in classes:
            print(f"  [간지] {c['label']} · {c['count']}명")
            print(f"   안내장 × {c['count']}장")
        print(f"\n총 {total + len(classes)}장 (안내장 {total} + 간지 {len(classes)})")
        return

    # 한글 OLE 기동
    log.info("한글 OLE 기동 중...")
    hwp = get_hwp()
    printer_label = PRINTER_NAME or "(기본 프린터)"
    duplex_label = {0: "단면", 1: "양면(긴면)", 2: "양면(짧은면)"}.get(DUPLEX, str(DUPLEX))
    log.info(f"안내장: {Path(hwpx_file).name}  |  프린터: {printer_label}  |  인쇄방식: {duplex_label}")
    print()

    tmp_files = []  # 인쇄 후 삭제할 임시 파일 목록

    try:
        for i, cls in enumerate(classes, 1):
            label = cls["label"]
            count = cls["count"]
            log.info(f"─── [{i}/{len(classes)}] {label} ({count}명) ───")

            # ① 간지 생성 및 인쇄
            log.info(f"  간지 생성 중...")
            sep_path = make_separator_hwpx(
                SEPARATOR_TEMPLATE,
                cls["grade"], cls["cls"], count
            )
            tmp_files.append(sep_path)

            log.info(f"  간지 인쇄 (1장 / 트레이 {SEPARATOR_TRAY})...")
            ok_sep = print_hwpx(hwp, sep_path, copies=1, printer=PRINTER_NAME,
                                duplex=0, paper_source=SEPARATOR_TRAY)
            if ok_sep:
                log.info(f"  ✓ 간지 전송 완료")
            else:
                log.warning(f"  ✗ 간지 전송 실패")

            log.info(f"  다음 안내장까지 {JOB_DELAY}초 대기...")
            time.sleep(JOB_DELAY)

            # ② 안내장 인쇄 (트레이 1 고정)
            MAIN_TRAY = 1
            log.info(f"  안내장 인쇄 ({count}장 / 트레이 {MAIN_TRAY})...")
            ok_main = print_hwpx(hwp, hwpx_file, copies=count, printer=PRINTER_NAME,
                                 duplex=DUPLEX, paper_source=MAIN_TRAY)
            if ok_main:
                log.info(f"  ✓ 안내장 전송 완료")
            else:
                log.warning(f"  ✗ 안내장 전송 실패")

            if i < len(classes):
                log.info(f"  다음 반까지 {JOB_DELAY}초 대기...")
                time.sleep(JOB_DELAY)
            print()

    finally:
        # 임시 간지 파일 삭제
        hwp.Quit()
        for f in tmp_files:
            try:
                Path(f).unlink()
            except Exception:
                pass

    log.info("=" * 50)
    log.info("인쇄 전송 완료!")
    log.info("프린터 출력물에서 간지를 기준으로 반별로 나누세요.")


if __name__ == "__main__":
    main()
